"""Excel 导入导出工具"""
from io import BytesIO
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_template(columns: List[dict]) -> BytesIO:
    """创建导入模板
    columns: [{"key": "username", "label": "用户名", "required": True}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin")
    header_border = Border(bottom=thin)

    headers = []
    for col in columns:
        label = col["label"]
        if col.get("required"):
            label += " *"
        headers.append(label)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = header_border

    # 备注行
    notes = [col.get("note", "") for col in columns]
    for col_idx, note in enumerate(notes, 1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font = Font(color="999999", size=9, italic=True)

    ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width = 20
    for col_idx in range(2, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_excel(headers: List[str], rows: List[List], sheet_name: str = "数据") -> BytesIO:
    """导出数据到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_import_file(file_bytes: bytes, expected_columns: List[str]) -> List[dict]:
    """解析导入文件，返回数据列表"""
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("文件为空")

    header_row = [str(h).replace(" *", "").strip() for h in rows[0]]
    for expected in expected_columns:
        if expected not in header_row:
            raise ValueError(f"缺少列: {expected}")

    col_indices = [header_row.index(expected) for expected in expected_columns]

    data = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {}
        for idx, col_idx in enumerate(col_indices):
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None:
                value = str(value).strip()
            record[expected_columns[idx]] = value if value else ""
        data.append(record)

    return data
